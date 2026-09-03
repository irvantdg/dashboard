"""Rute administrasi: manajemen user, audit log, import CSV/Excel dengan staging."""
import csv
import io
import re
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, EmailStr

from auth import get_current_user, hash_password, require_roles
from common import audit, db, doc_out, docs_out, new_id, now_utc
from metrics import parse_ym

router = APIRouter(prefix="/api", tags=["admin"])

admin = require_roles("admin")

IMPORT_FIELDS = {
    "member": [
        {"field": "member_code", "label": "Kode Member", "required": True},
        {"field": "member_name", "label": "Nama Member", "required": True},
        {"field": "alias", "label": "Alias", "required": False},
        {"field": "member_type", "label": "Tipe Member", "required": False},
        {"field": "pic", "label": "PIC", "required": False},
        {"field": "status", "label": "Status", "required": False},
    ],
    "product": [
        {"field": "product_code", "label": "Kode Produk", "required": True},
        {"field": "product_name", "label": "Nama Produk", "required": True},
        {"field": "category", "label": "Kategori", "required": False},
        {"field": "description", "label": "Deskripsi", "required": False},
        {"field": "status", "label": "Status", "required": False},
    ],
    "matrix": [
        {"field": "member_code", "label": "Kode Member", "required": True},
        {"field": "product_code", "label": "Kode Produk", "required": True},
        {"field": "position", "label": "Posisi", "required": True},
        {"field": "status", "label": "Status Implementasi", "required": True},
        {"field": "status_date", "label": "Tanggal Status (YYYY-MM)", "required": False},
        {"field": "pic", "label": "PIC", "required": False},
        {"field": "notes", "label": "Catatan", "required": False},
    ],
    "transaction": [
        {"field": "period", "label": "Periode (YYYY-MM)", "required": True},
        {"field": "member_code", "label": "Kode Member", "required": True},
        {"field": "product_code", "label": "Kode Produk", "required": True},
        {"field": "position", "label": "Posisi (Issuer/Acquirer)", "required": True},
        {"field": "aggregation_type", "label": "Tipe Agregasi (Single Side/Cross)", "required": True},
        {"field": "volume", "label": "Volume", "required": True},
        {"field": "nominal", "label": "Nominal", "required": True},
        {"field": "fee", "label": "Fee", "required": False},
    ],
}

MATRIX_STATUSES = ["Live", "UAT", "Development", "Preparation", "On Hold", "Not Implemented"]


@router.get("/imports/fields")
async def import_fields(user: dict = Depends(admin)):
    return IMPORT_FIELDS


def _parse_file(filename: str, content: bytes):
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = [dict(r) for r in reader]
        return reader.fieldnames or [], rows
    if filename.lower().endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        if not data:
            return [], []
        headers = [str(h).strip() if h is not None else f"kolom_{i+1}" for i, h in enumerate(data[0])]
        rows = [{headers[i]: ("" if v is None else str(v)) for i, v in enumerate(r)} for r in data[1:]]
        return headers, rows
    raise HTTPException(422, "Format file tidak didukung. Gunakan CSV atau XLSX.")


@router.post("/imports")
async def upload_import(import_type: str, file: UploadFile, user: dict = Depends(admin)):
    if import_type not in IMPORT_FIELDS:
        raise HTTPException(422, "Tipe import tidak valid")
    if file.size and file.size > 10 * 1024 * 1024:
        raise HTTPException(422, "Ukuran file maksimal 10 MB")
    content = await file.read()
    headers, rows = _parse_file(file.filename, content)
    if not rows:
        raise HTTPException(422, "File kosong atau tidak memiliki baris data")
    batch_id = new_id()
    await db.import_staging.insert_one({"_id": batch_id, "rows": rows[:5000]})
    batch = {"_id": batch_id, "import_type": import_type, "file_name": file.filename,
             "status": "staged", "total_rows": len(rows), "accepted_rows": 0, "rejected_rows": 0,
             "errors": [], "uploaded_by": user["email"], "uploaded_at": now_utc(), "headers": headers}
    await db.import_batches.insert_one(batch)
    await audit(user, "upload_file", "import_batch", batch_id, new={"file": file.filename, "type": import_type, "rows": len(rows)})
    # Saran pemetaan otomatis berdasarkan kemiripan nama kolom
    suggested = {}
    for fdef in IMPORT_FIELDS[import_type]:
        for h in headers:
            if h.strip().lower() in (fdef["field"].lower(), fdef["label"].lower()):
                suggested[fdef["field"]] = h
                break
    return {"batch_id": batch_id, "headers": headers, "total_rows": len(rows),
            "preview": rows[:15], "suggested_mapping": suggested}


class ValidateBody(BaseModel):
    mapping: dict  # {field: source_column}


async def _validate(import_type: str, rows, mapping):
    member_codes = {m["member_code"]: m["member_name"] for m in await db.members.find({}, {"member_code": 1, "member_name": 1, "_id": 0}).to_list(1000)}
    product_codes = {p["product_code"] for p in await db.products.find({}, {"product_code": 1, "_id": 0}).to_list(1000)}
    errors, accepted, rejected = [], [], []
    seen = set()

    def get(row, field):
        col = mapping.get(field)
        return (row.get(col) or "").strip() if col else ""

    for idx, row in enumerate(rows, start=2):
        errs = []
        for fdef in IMPORT_FIELDS[import_type]:
            if fdef["required"] and not get(row, fdef["field"]):
                errs.append({"field": fdef["field"], "message": f"Kolom wajib '{fdef['label']}' kosong", "level": "error"})
        mc, pc = get(row, "member_code"), get(row, "product_code")
        if import_type in ("matrix", "transaction"):
            if mc and mc not in member_codes:
                errs.append({"field": "member_code", "message": f"Kode member '{mc}' tidak dikenal", "level": "error"})
            if pc and pc not in product_codes:
                errs.append({"field": "product_code", "message": f"Kode produk '{pc}' tidak dikenal", "level": "error"})
        if import_type == "matrix":
            if get(row, "position") and get(row, "position") not in ("Issuer", "Acquirer", "Issuer & Acquirer"):
                errs.append({"field": "position", "message": "Posisi tidak valid", "level": "error"})
            if get(row, "status") and get(row, "status") not in MATRIX_STATUSES:
                errs.append({"field": "status", "message": "Status implementasi tidak valid", "level": "error"})
        if import_type == "transaction":
            per = get(row, "period")
            if per and not re.match(r"^\d{4}-(0[1-9]|1[0-2])$", per):
                errs.append({"field": "period", "message": f"Periode '{per}' tidak valid (format YYYY-MM)", "level": "error"})
            if get(row, "position") and get(row, "position") not in ("Issuer", "Acquirer"):
                errs.append({"field": "position", "message": "Posisi harus Issuer atau Acquirer", "level": "error"})
            if get(row, "aggregation_type") and get(row, "aggregation_type") not in ("Single Side", "Cross"):
                errs.append({"field": "aggregation_type", "message": "Tipe agregasi harus Single Side atau Cross", "level": "error"})
            for nf in ("volume", "nominal", "fee"):
                v = get(row, nf)
                if v:
                    try:
                        fv = _num(v)
                        if nf == "volume" and fv < 0:
                            errs.append({"field": nf, "message": "Volume tidak boleh negatif", "level": "error"})
                    except ValueError:
                        errs.append({"field": nf, "message": f"Nilai numerik '{v}' tidak valid", "level": "error"})
            if mc and get(row, "member_name") and member_codes.get(mc) and get(row, "member_name") != member_codes[mc]:
                errs.append({"field": "member_name", "message": "Nama member tidak konsisten dengan kode member", "level": "warning"})
            key = (per, mc, pc, get(row, "position"), get(row, "aggregation_type"))
        elif import_type == "matrix":
            key = (mc, pc)
        else:
            key = (mc,) if import_type == "member" else (pc,)
        if key in seen:
            errs.append({"field": "-", "message": "Baris duplikat di dalam file", "level": "error"})
        seen.add(key)
        has_error = any(e["level"] == "error" for e in errs)
        for e in errs:
            e["row"] = idx
        errors.extend(errs)
        (rejected if has_error else accepted).append(idx - 2)
    return {"errors": errors[:500], "accepted": accepted, "rejected": rejected}


@router.post("/imports/{batch_id}/validate")
async def validate_import(batch_id: str, body: ValidateBody, user: dict = Depends(admin)):
    batch = await db.import_batches.find_one({"_id": batch_id})
    if not batch:
        raise HTTPException(404, "Batch import tidak ditemukan")
    if batch["status"] == "committed":
        raise HTTPException(409, "Batch sudah di-commit")
    staging = await db.import_staging.find_one({"_id": batch_id})
    result = await _validate(batch["import_type"], staging["rows"], body.mapping)
    await db.import_batches.update_one({"_id": batch_id}, {"$set": {
        "status": "validated", "mapping": body.mapping,
        "accepted_rows": len(result["accepted"]), "rejected_rows": len(result["rejected"]),
        "errors": result["errors"]}})
    await db.import_staging.update_one({"_id": batch_id}, {"$set": {"accepted_idx": result["accepted"]}})
    return {"accepted": len(result["accepted"]), "rejected": len(result["rejected"]),
            "errors": result["errors"][:200]}


def _num(s):
    s = (s or "").strip()
    if not s:
        return 0
    if re.match(r"^\d{1,3}(\.\d{3})+$", s):
        s = s.replace(".", "")
    return float(s.replace(",", ""))


@router.post("/imports/{batch_id}/commit")
async def commit_import(batch_id: str, user: dict = Depends(admin)):
    batch = await db.import_batches.find_one({"_id": batch_id})
    if not batch or batch["status"] != "validated":
        raise HTTPException(409, "Batch belum divalidasi atau sudah di-commit")
    staging = await db.import_staging.find_one({"_id": batch_id})
    mapping = batch["mapping"]
    rows = staging["rows"]
    accepted_idx = set(staging.get("accepted_idx", []))
    itype = batch["import_type"]

    def get(row, field):
        col = mapping.get(field)
        return (row.get(col) or "").strip() if col else ""

    n = 0
    for i in sorted(accepted_idx):
        row = rows[i]
        if itype == "member":
            await db.members.update_one({"member_code": get(row, "member_code")}, {"$set": {
                "member_code": get(row, "member_code"), "member_name": get(row, "member_name"),
                "alias": get(row, "alias") or get(row, "member_code"),
                "member_type": get(row, "member_type") or "Bank Umum",
                "pic": get(row, "pic"), "status": get(row, "status") or "Aktif"}}, upsert=True)
        elif itype == "product":
            await db.products.update_one({"product_code": get(row, "product_code")}, {"$set": {
                "product_code": get(row, "product_code"), "product_name": get(row, "product_name"),
                "category": get(row, "category") or "Lainnya",
                "description": get(row, "description"),
                "is_cross_border": False,
                "status": get(row, "status") or "Aktif"}}, upsert=True)
        elif itype == "matrix":
            m = await db.members.find_one({"member_code": get(row, "member_code")})
            p = await db.products.find_one({"product_code": get(row, "product_code")})
            sd = parse_ym(get(row, "status_date")) if get(row, "status_date") else now_utc()
            await db.member_products.update_one(
                {"member_code": get(row, "member_code"), "product_code": get(row, "product_code")},
                {"$set": {"member_code": get(row, "member_code"), "member_name": m["member_name"] if m else "",
                          "member_alias": m.get("alias", "") if m else "",
                          "product_code": get(row, "product_code"), "product_name": p["product_name"] if p else "",
                          "position": get(row, "position"), "status": get(row, "status"),
                          "status_date": sd, "pic": get(row, "pic"), "notes": get(row, "notes"),
                          "updated_by": user["email"], "updated_at": now_utc()}}, upsert=True)
        elif itype == "transaction":
            m = await db.members.find_one({"member_code": get(row, "member_code")})
            p = await db.products.find_one({"product_code": get(row, "product_code")})
            per = get(row, "period")
            await db.transactions.update_one({
                "period_ym": per, "member_code": get(row, "member_code"),
                "product_code": get(row, "product_code"), "position": get(row, "position"),
                "aggregation_type": get(row, "aggregation_type")}, {"$set": {
                "period": parse_ym(per), "period_ym": per,
                "member_code": get(row, "member_code"),
                "member_name": m["member_name"] if m else get(row, "member_code"),
                "member_alias": m.get("alias", "") if m else "",
                "product_code": get(row, "product_code"),
                "product_name": p["product_name"] if p else get(row, "product_code"),
                "product_category": p.get("category", "Lainnya") if p else "Lainnya",
                "product_border": p.get("is_cross_border", False) if p else False,
                "position": get(row, "position"), "aggregation_type": get(row, "aggregation_type"),
                "volume": int(_num(get(row, "volume"))), "nominal": _num(get(row, "nominal")),
                "fee": _num(get(row, "fee")),
                "batch_id": batch_id, "updated_at": now_utc()}}, upsert=True)
        n += 1
    await db.import_batches.update_one({"_id": batch_id}, {"$set": {"status": "committed", "committed_at": now_utc(), "committed_by": user["email"]}})
    await db.import_staging.delete_one({"_id": batch_id})
    await audit(user, "commit_import", "import_batch", batch_id, new={"accepted": n, "type": itype})
    return {"message": f"Import selesai: {n} baris diterima", "committed_rows": n}


@router.get("/imports")
async def list_imports(user: dict = Depends(admin)):
    rows = await db.import_batches.find({}, {"errors": 0}).sort("uploaded_at", -1).limit(100).to_list(None)
    return {"batches": docs_out(rows)}


@router.get("/imports/{batch_id}")
async def import_detail(batch_id: str, user: dict = Depends(admin)):
    b = await db.import_batches.find_one({"_id": batch_id})
    if not b:
        raise HTTPException(404, "Batch tidak ditemukan")
    return doc_out(b)


# ---------- Manajemen User ----------

@router.get("/users")
async def list_users(user: dict = Depends(admin)):
    rows = await db.users.find({}, {"password_hash": 0}).to_list(500)
    return {"users": docs_out(rows)}


class UserCreate(BaseModel):
    name: str
    email: EmailStr
    password: str
    role: str


class UserUpdate(BaseModel):
    name: str | None = None
    role: str | None = None
    status: str | None = None
    password: str | None = None
    view_fee: bool | None = None


@router.post("/users")
async def create_user(body: UserCreate, user: dict = Depends(admin)):
    from auth import RegisterBody, register
    return await register(RegisterBody(name=body.name, email=body.email, password=body.password, role=body.role), None, user)


@router.put("/users/{uid}")
async def update_user(uid: str, body: UserUpdate, user: dict = Depends(admin)):
    target = await db.users.find_one({"_id": uid})
    if not target:
        raise HTTPException(404, "Pengguna tidak ditemukan")
    upd, prev = {}, {}
    if body.name:
        upd["name"] = body.name.strip()
    if body.role:
        if body.role not in ("admin", "analyst", "management"):
            raise HTTPException(422, "Peran tidak valid")
        upd["role"] = body.role
    if body.status:
        upd["status"] = body.status
    if body.view_fee is not None:
        upd["permissions.view_fee"] = body.view_fee
    if body.password:
        if len(body.password) < 8:
            raise HTTPException(422, "Kata sandi minimal 8 karakter")
        upd["password_hash"] = hash_password(body.password)
        upd["token_version"] = target.get("token_version", 0) + 1
    if not upd:
        return {"message": "Tidak ada perubahan"}
    for k in ("role", "status"):
        if k in upd:
            prev[k] = target.get(k)
    prev["view_fee"] = target.get("permissions", {}).get("view_fee", True)
    await db.users.update_one({"_id": uid}, {"$set": upd})
    await audit(user, "update_user", "user", uid, prev=prev,
                new={k: v for k, v in upd.items() if k != "password_hash"})
    return {"message": "Pengguna diperbarui"}


# ---------- Audit Log ----------

@router.get("/audit")
async def audit_log(page: int = 1, page_size: int = 50, user: dict = Depends(admin)):
    total = await db.audit_logs.count_documents({})
    rows = await db.audit_logs.find().sort("timestamp", -1).skip((page - 1) * page_size).limit(page_size).to_list(None)
    return {"rows": docs_out(rows), "total": total, "page": page, "page_size": page_size}
